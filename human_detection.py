# Code adapted from Tensorflow Object Detection Framework and madhawav
# https://github.com/tensorflow/models/blob/master/research/object_detection/object_detection_tutorial.ipynb
# Tensorflow Object Detection Detector


import numpy as np
import tensorflow as tf
import cv2
import time
import random
# import xlwt
import time
import os
import sys
from openpyxl import Workbook


class DetectorAPI:
    def __init__(self, path_to_ckpt):
        self.path_to_ckpt = path_to_ckpt

        self.detection_graph = tf.Graph()
        with self.detection_graph.as_default():
            od_graph_def = tf.GraphDef()
            with tf.gfile.GFile(self.path_to_ckpt, 'rb') as fid:
                serialized_graph = fid.read()
                od_graph_def.ParseFromString(serialized_graph)
                tf.import_graph_def(od_graph_def, name='')

        self.default_graph = self.detection_graph.as_default()
        self.sess = tf.Session(graph=self.detection_graph)

        # Definite input and output Tensors for detection_graph
        self.image_tensor = self.detection_graph.get_tensor_by_name('image_tensor:0')
        # Each box represents a part of the image where a particular object was detected.
        self.detection_boxes = self.detection_graph.get_tensor_by_name('detection_boxes:0')
        # Each score represent how level of confidence for each of the objects.
        # Score is shown on the result image, together with the class label.
        self.detection_scores = self.detection_graph.get_tensor_by_name('detection_scores:0')
        self.detection_classes = self.detection_graph.get_tensor_by_name('detection_classes:0')
        self.num_detections = self.detection_graph.get_tensor_by_name('num_detections:0')

    def processFrame(self, image, frames, count):
        # Expand dimensions since the trained_model expects images to have shape: [1, None, None, 3]
        image_np_expanded = np.expand_dims(image, axis=0)
        # Actual detection.
        start_time = time.time()
        (boxes, scores, classes, num) = self.sess.run(
            [self.detection_boxes, self.detection_scores, self.detection_classes, self.num_detections],
            feed_dict={self.image_tensor: image_np_expanded})
        end_time = time.time()
        
        time_elapsed = end_time - start_time
        percent_complete = int(100*count/frames)
        time_left = (1-percent_complete/100)*frames*time_elapsed/60
        print("Elapsed Time: {:.5}    Minutes remaining: {:.5}".format(time_elapsed, time_left))
        print("Frame: ", count, r"/", frames, " ", percent_complete, "%") 

        im_height, im_width,_ = image.shape
        boxes_list = [None for i in range(boxes.shape[1])]
        for i in range(boxes.shape[1]):
            boxes_list[i] = (int(boxes[0,i,0] * im_height),
                        int(boxes[0,i,1]*im_width),
                        int(boxes[0,i,2] * im_height),
                        int(boxes[0,i,3]*im_width))

        return boxes_list, scores[0].tolist(), [int(x) for x in classes[0].tolist()], int(num[0])

    def close(self):
        self.sess.close()
        self.default_graph.close()

if __name__ == "__main__":
    # Model taken from https://github.com/tensorflow/models/blob/master/research/object_detection/g3doc/detection_model_zoo.md
    model_path = r'/home/sam/841project/faster_rcnn_nas_coco_2018_01_28/frozen_inference_graph.pb'
    odapi = DetectorAPI(path_to_ckpt=model_path)
    threshold = 0.4
    path = os.path.join(r"/home/sam/841project/dev/video_input", sys.argv[1]+'.mp4')
    cap = cv2.VideoCapture(path)
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)   # float
    height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT) # float
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    # fps = cap.get(7)
    fps = 30
    print('Width: %s  Height: %s  FPS: %s' % (width, height, fps))
    path = os.path.join(r'video_output', sys.argv[1]+'_rected.mp4')
    out = cv2.VideoWriter(path, fourcc, int(fps), (int(width),int(height)))
    count = 0
    column_count = 0
#     # Generate some random colors
#     num_colors = 40
#     colors = [0]*num_colors
#     for i in range(len(colors)):
#         colors[i] = (random.randint(0,255), random.randint(0,255), random.randint(0,255))

#     book = xlwt.Workbook()
#     sheet = book.add_sheet('positions')
    wb = Workbook()
    ws = wb.active
    
    while True:
        count = count + 1
        ret, frame = cap.read()
        if not ret:
            break
            
#         frame = cv2.resize(frame, (1280, 720))
        boxes, scores, classes, num = odapi.processFrame(frame, total_frames, count)
#         print(type(boxes), len(boxes), boxes[0], "scores:", scores[1])
        # Visualization of the results of a detection.
        for i in range(len(boxes)):
            # Class 1 represents human
            if classes[i] == 1 and scores[i] > threshold:
                box = boxes[i]
                cv2.rectangle(frame,(box[1],box[0]),(box[3],box[2]),(0,255,0),1)
                box_string = "(%s, %s, %s, %s)" % (box[1],box[0],box[3],box[2])
                ws.cell(row=count+1, column=i+1, value=box_string)
#                 sheet.write(count, i, box_string)
        
        out.write(frame)
        
#         cv2.imshow("preview", img)
#         key = cv2.waitKey(1)
#         if key & 0xFF == ord('q'):
#             break
    path = os.path.join(r'video_output', sys.argv[1]+'_rected.xlsx')
    wb.save(path)
#     book.save(path)
    cap.release()
    out.release()
